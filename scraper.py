# -*- coding: utf-8 -*-
"""
今日头条作者文章爬虫 (Toutiao Author Article Scraper)
=====================================================
功能特点：
1. 全自动化：基于 Patchright / Playwright 自动绕过头条反爬、JS混淆与动态签名机制。
2. 完整字段：爬取文章标题、正文、发布时间、阅读量、点赞量、评论量、头条原链接、配图等。
3. 结构化排版：智能解析 HTML 并转换为干净优雅的 Markdown 格式（含 YAML Frontmatter 元数据）。
4. 多格式导出：自动生成 Markdown 单篇文件、Excel 表格汇总、CSV 表格以及 JSON 完整数据。
5. 断点续爬：已下载文章自动跳过，支持海量文章增量抓取与防封延迟控制。
"""

import os
import re
import sys
import time
import json
import asyncio
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from urllib.parse import urlparse, parse_qs
import urllib.request

# 解决 Windows 控制台编码问题
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# 浏览器内核驱动
try:
    from patchright.async_api import async_playwright, Browser, BrowserContext, Page
except ImportError:
    from playwright.async_api import async_playwright, Browser, BrowserContext, Page

from bs4 import BeautifulSoup, NavigableString, Tag
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from rich.console import Console
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeRemainingColumn
from rich.table import Table
from rich.panel import Panel

console = Console(force_terminal=True, legacy_windows=False)

def sanitize_filename(name: str, max_length: int = 60) -> str:
    """去除文件名中的非法字符并限制最大长度"""
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_length:
        name = name[:max_length]
    return name or 'untitled'

def format_timestamp(ts) -> str:
    """将时间戳转换为格式化日期时间"""
    if not ts:
        return ''
    try:
        ts = int(ts)
        if ts > 10000000000: # 毫秒戳
            ts = ts // 1000
        return datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return str(ts)

def format_date_for_file(ts) -> str:
    """将时间戳转换为年月日用于文件名前缀"""
    if not ts:
        return datetime.now().strftime('%Y%m%d')
    try:
        ts = int(ts)
        if ts > 10000000000:
            ts = ts // 1000
        return datetime.fromtimestamp(ts).strftime('%Y%m%d')
    except Exception:
        return datetime.now().strftime('%Y%m%d')

def html_to_markdown(soup_or_tag: Tag, local_image_map: Optional[Dict[str, str]] = None) -> str:
    """将文章的 BeautifulSoup 节点转换为优雅的 Markdown 格式"""
    if not soup_or_tag:
        return ''

    def process_node(node):
        if isinstance(node, NavigableString):
            return str(node)
        if not isinstance(node, Tag):
            return ''

        tag_name = node.name.lower()

        # 过滤无用节点
        if tag_name in ['script', 'style', 'button', 'noscript', 'iframe', 'svg']:
            return ''

        # 标题标签
        if tag_name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            level = int(tag_name[1])
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n{'#' * level} {inner}\n\n" if inner else ''

        # 段落
        elif tag_name == 'p':
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n{inner}\n\n" if inner else ''

        # 加粗
        elif tag_name in ['strong', 'b']:
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"**{inner}**" if inner else ''

        # 斜体
        elif tag_name in ['em', 'i']:
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"*{inner}*" if inner else ''

        # 图片
        elif tag_name == 'img':
            src = node.get('data-src') or node.get('src') or ''
            if src.startswith('data:'):
                src = node.get('data-src') or ''
            alt = node.get('alt') or '配图'
            if src:
                final_src = local_image_map.get(src, src) if local_image_map else src
                return f"\n\n![{alt}]({final_src})\n\n"
            return ''

        # 引用块
        elif tag_name == 'blockquote':
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n> {inner}\n\n" if inner else ''

        # 列表
        elif tag_name in ['ul', 'ol']:
            items = []
            for i, li in enumerate(node.find_all('li', recursive=False)):
                li_text = ''.join(process_node(c) for c in li.children).strip()
                prefix = f"{i+1}. " if tag_name == 'ol' else "- "
                items.append(f"{prefix}{li_text}")
            return '\n\n' + '\n'.join(items) + '\n\n'

        # 换行
        elif tag_name == 'br':
            return '\n'

        # 链接
        elif tag_name == 'a':
            inner = ''.join(process_node(c) for c in node.children).strip()
            href = node.get('href', '')
            if href and inner:
                return f"[{inner}]({href})"
            return inner

        else:
            return ''.join(process_node(c) for c in node.children)

    raw_md = process_node(soup_or_tag)
    cleaned = re.sub(r'\n{3,}', '\n\n', raw_md).strip()
    return cleaned


class ToutiaoSpider:
    """今日头条作者文章爬虫"""

    DEFAULT_USER_AGENT = (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/128.0.0.0 Safari/537.36'
    )

    def __init__(
        self,
        author_url: str,
        output_dir: str = './toutiao_output',
        max_articles: Optional[int] = None,
        fetch_content: bool = True,
        download_images: bool = False,
        headless: bool = True,
        delay: float = 0.8,
        scroll_delay: float = 1.2,
        max_scroll_retries: int = 5
    ):
        self.author_url = author_url
        self.output_dir = Path(output_dir)
        self.max_articles = max_articles
        self.fetch_content = fetch_content
        self.download_images = download_images
        self.headless = headless
        self.delay = delay
        self.scroll_delay = scroll_delay
        self.max_scroll_retries = max_scroll_retries

        # 目录初始化
        self.articles_dir = self.output_dir / 'articles'
        self.images_dir = self.output_dir / 'images'
        self.articles_dir.mkdir(parents=True, exist_ok=True)
        if self.download_images:
            self.images_dir.mkdir(parents=True, exist_ok=True)

        self.author_info: Dict[str, Any] = {}
        self.articles_map: Dict[str, Dict[str, Any]] = {}

    async def _init_browser(self, p):
        """初始化带反爬隐蔽特性的浏览器上下文"""
        browser = await p.chromium.launch(
            headless=self.headless,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-dev-shm-usage',
            ]
        )
        context = await browser.new_context(
            user_agent=self.DEFAULT_USER_AGENT,
            viewport={'width': 1366, 'height': 850},
            locale='zh-CN'
        )
        # 注入防自动化检测脚本
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)
        return browser, context

    async def crawl_author_feed(self, page: Page) -> List[Dict[str, Any]]:
        """
        滚动抓取作者文章列表接口 (/api/pc/list/user/feed)
        """
        console.print(Panel.fit(f"[bold cyan]正在访问作者主页:[/bold cyan] [underline]{self.author_url}[/underline]"))

        has_more_flag = True
        captured_batches = 0

        async def on_response(response):
            nonlocal has_more_flag, captured_batches
            if '/api/pc/list/user/feed' in response.url:
                try:
                    data = await response.json()
                    has_more = data.get('has_more', False)
                    items = data.get('data', [])
                    captured_batches += 1
                    
                    for item in items:
                        gid = str(item.get('group_id') or item.get('item_id') or '')
                        if not gid:
                            continue
                        if gid not in self.articles_map:
                            title = (item.get('title') or item.get('abstract') or '无标题').strip()
                            pub_time = item.get('publish_time') or item.get('behot_time')
                            images = []
                            for img in item.get('image_list', []):
                                if isinstance(img, dict) and 'url' in img:
                                    images.append(img['url'])

                            self.articles_map[gid] = {
                                'group_id': gid,
                                'title': title,
                                'abstract': item.get('abstract', ''),
                                'publish_timestamp': pub_time,
                                'publish_time': format_timestamp(pub_time),
                                'read_count': item.get('read_count', 0),
                                'digg_count': item.get('digg_count', 0),
                                'comment_count': item.get('comment_count') or item.get('comments_count', 0),
                                'article_url': f"https://www.toutiao.com/article/{gid}/",
                                'share_url': item.get('share_url', ''),
                                'image_list': images,
                                'has_video': item.get('has_video', False),
                                'content_markdown': '',
                                'local_file': '',
                                'status': 'pending'
                            }

                    if not has_more:
                        has_more_flag = False
                except Exception:
                    pass

        page.on('response', on_response)

        # 打开页面
        await page.goto(self.author_url, wait_until='domcontentloaded', timeout=30000)
        await asyncio.sleep(2)

        # 尝试提取作者基础信息
        try:
            name_el = await page.query_selector('.user-info .name, .name, [class*="user-name"]')
            if name_el:
                self.author_info['name'] = (await name_el.inner_text()).strip()
            desc_el = await page.query_selector('.user-info .desc, .desc, [class*="user-desc"]')
            if desc_el:
                self.author_info['desc'] = (await desc_el.inner_text()).strip()
        except Exception:
            pass

        console.print(f"[green][OK] 作者主页加载成功: {self.author_info.get('name', '今日头条作者')}[/green]")

        # 动态滚动加载更多文章
        last_count = 0
        same_count_streak = 0

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[bold yellow]{task.fields[articles_count]} 篇[/bold yellow]"),
            console=console
        ) as progress:
            task = progress.add_task("[cyan]正在滚动加载文章列表...", total=self.max_articles or 1000, articles_count=0)

            while has_more_flag:
                await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                await asyncio.sleep(self.scroll_delay)

                current_count = len(self.articles_map)
                progress.update(task, completed=min(current_count, self.max_articles or 1000), articles_count=current_count)

                # 达到用户指定的上限
                if self.max_articles and current_count >= self.max_articles:
                    console.print(f"[yellow]已达到设定的最大抓取数量 ({self.max_articles} 篇)[/yellow]")
                    break

                # 检查是否停止加载
                if current_count == last_count:
                    same_count_streak += 1
                    if same_count_streak >= self.max_scroll_retries:
                        await page.evaluate('window.scrollBy(0, -300)')
                        await asyncio.sleep(0.5)
                        await page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        await asyncio.sleep(self.scroll_delay)
                        if len(self.articles_map) == last_count:
                            console.print("[dim]已滚动至页面底部，所有文章列表加载完毕。[/dim]")
                            break
                else:
                    same_count_streak = 0
                    last_count = current_count

        article_list = list(self.articles_map.values())
        if self.max_articles:
            article_list = article_list[:self.max_articles]

        console.print(f"[bold green][OK] 成功检索到 {len(article_list)} 篇有效文章！[/bold green]")
        return article_list

    def _download_image(self, img_url: str, save_path: Path) -> bool:
        """下载图片到本地"""
        try:
            req = urllib.request.Request(
                img_url,
                headers={
                    'User-Agent': self.DEFAULT_USER_AGENT,
                    'Referer': 'https://www.toutiao.com/'
                }
            )
            with urllib.request.urlopen(req, timeout=10) as resp, open(save_path, 'wb') as f:
                f.write(resp.read())
            return True
        except Exception:
            return False

    async def crawl_article_content(self, page: Page, item: Dict[str, Any]) -> Dict[str, Any]:
        """
        进入单篇文章详情页，提取完整正文、排版并转为 Markdown
        """
        gid = item['group_id']
        url = item['article_url']
        pub_date = format_date_for_file(item['publish_timestamp'])
        safe_title = sanitize_filename(item['title'])
        filename = f"{pub_date}_{gid}_{safe_title}.md"
        filepath = self.articles_dir / filename

        # 检查是否已存在（断点续爬）
        if filepath.exists() and filepath.stat().st_size > 100:
            item['local_file'] = str(filepath.relative_to(self.output_dir))
            item['status'] = 'cached'
            return item

        try:
            await page.goto(url, wait_until='domcontentloaded', timeout=20000)

            # 等待正文渲染
            article_node = None
            for selector in ['article', '.article-content', '.tt-article-content', '.s-content', '.main-content']:
                try:
                    await page.wait_for_selector(selector, timeout=4000)
                    article_node = selector
                    break
                except Exception:
                    continue

            if not article_node:
                # 尝试直接读取 body
                html = await page.content()
                soup = BeautifulSoup(html, 'html.parser')
                article_tag = soup.find('article') or soup.find('div', class_='article-content')
            else:
                article_html = await page.inner_html(article_node)
                soup = BeautifulSoup(article_html, 'html.parser')
                article_tag = soup

            # 抓取页面上的标题
            try:
                h1_el = await page.query_selector('h1')
                if h1_el:
                    page_title = (await h1_el.inner_text()).strip()
                    if page_title:
                        item['title'] = page_title
            except Exception:
                pass

            # 图片下载与本地映射处理
            local_image_map = {}
            if self.download_images and article_tag:
                img_nodes = article_tag.find_all('img')
                if img_nodes:
                    article_img_dir = self.images_dir / gid
                    article_img_dir.mkdir(parents=True, exist_ok=True)
                    for idx, img_tag in enumerate(img_nodes, 1):
                        src = img_tag.get('data-src') or img_tag.get('src') or ''
                        if src and not src.startswith('data:'):
                            ext = 'jpg'
                            if '.png' in src: ext = 'png'
                            elif '.webp' in src: ext = 'webp'
                            elif '.gif' in src: ext = 'gif'
                            img_filename = f"{idx:02d}.{ext}"
                            img_save_path = article_img_dir / img_filename
                            if not img_save_path.exists():
                                self._download_image(src, img_save_path)
                            rel_img_path = f"../images/{gid}/{img_filename}"
                            local_image_map[src] = rel_img_path

            # 转换为 Markdown
            content_md = html_to_markdown(article_tag, local_image_map)
            item['content_markdown'] = content_md

            # 生成带 YAML Frontmatter 的完整 Markdown 文件
            author_name = self.author_info.get('name', '今日头条作者')
            md_full_text = (
                f"---\n"
                f"title: \"{item['title']}\"\n"
                f"author: \"{author_name}\"\n"
                f"publish_time: \"{item['publish_time']}\"\n"
                f"group_id: \"{gid}\"\n"
                f"read_count: {item.get('read_count', 0)}\n"
                f"digg_count: {item.get('digg_count', 0)}\n"
                f"comment_count: {item.get('comment_count', 0)}\n"
                f"source_url: \"{url}\"\n"
                f"---\n\n"
                f"# {item['title']}\n\n"
                f"> **发布时间**：{item['publish_time']} | **阅读量**：{item.get('read_count', 0)} | **点赞**：{item.get('digg_count', 0)} | **原文链接**：[{url}]({url})\n\n"
                f"---\n\n"
                f"{content_md}\n"
            )

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(md_full_text)

            item['local_file'] = str(filepath.relative_to(self.output_dir))
            item['status'] = 'downloaded'

        except Exception as e:
            item['status'] = f"failed: {str(e)}"

        return item

    def export_excel(self, articles: List[Dict[str, Any]]) -> str:
        """导出美化的 Excel 汇总表格"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文章汇总"

        # 表头样式
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Microsoft YaHei", size=10)
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        headers = ["序号", "文章ID", "发布时间", "文章标题", "阅读量", "点赞量", "评论量", "文章摘要", "原文链接", "本地Markdown文件"]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        for idx, item in enumerate(articles, 1):
            row_data = [
                idx,
                item.get('group_id', ''),
                item.get('publish_time', ''),
                item.get('title', ''),
                item.get('read_count', 0),
                item.get('digg_count', 0),
                item.get('comment_count', 0),
                item.get('abstract', ''),
                item.get('article_url', ''),
                item.get('local_file', '')
            ]
            ws.append(row_data)
            row_num = idx + 1
            ws.row_dimensions[row_num].height = 22

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = border
                if col_idx in [1, 2, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [3, 4, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 调整列宽
        col_widths = {1: 8, 2: 24, 3: 20, 4: 45, 5: 12, 6: 12, 7: 12, 8: 40, 9: 35, 10: 35}
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        excel_path = self.output_dir / 'articles_summary.xlsx'
        wb.save(excel_path)
        return str(excel_path)

    def export_csv(self, articles: List[Dict[str, Any]]) -> str:
        """导出 CSV 汇总表格 (带 UTF-8 BOM，防止 Excel 打开乱码)"""
        import csv
        csv_path = self.output_dir / 'articles_summary.csv'
        headers = ["序号", "文章ID", "发布时间", "文章标题", "阅读量", "点赞量", "评论量", "摘要", "原文链接", "本地文件"]
        
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for idx, item in enumerate(articles, 1):
                writer.writerow([
                    idx,
                    item.get('group_id', ''),
                    item.get('publish_time', ''),
                    item.get('title', ''),
                    item.get('read_count', 0),
                    item.get('digg_count', 0),
                    item.get('comment_count', 0),
                    item.get('abstract', ''),
                    item.get('article_url', ''),
                    item.get('local_file', '')
                ])
        return str(csv_path)

    def export_json(self, articles: List[Dict[str, Any]]) -> str:
        """导出完整 JSON 数据"""
        json_path = self.output_dir / 'articles_data.json'
        payload = {
            'crawler_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'author_info': self.author_info,
            'total_articles': len(articles),
            'articles': articles
        }
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return str(json_path)

    async def run(self):
        """主执行流程"""
        start_time = time.time()
        console.print(Panel.fit("[bold green]今日头条作者文章爬虫启动[/bold green]\n"
                                f"目标作者: {self.author_url}\n"
                                f"输出目录: {self.output_dir.resolve()}\n"
                                f"抓取正文: {'是' if self.fetch_content else '否 (仅抓元数据)'}\n"
                                f"最大篇数: {self.max_articles or '全部'}"))

        async with async_playwright() as p:
            browser, context = await self._init_browser(p)
            page = await context.new_page()

            try:
                # 第一步：抓取文章列表
                articles = await self.crawl_author_feed(page)
                if not articles:
                    console.print("[bold red]未获取到任何文章，请检查链接或网络状态。[/bold red]")
                    return

                # 第二步：抓取正文内容
                if self.fetch_content:
                    console.print(f"\n[bold cyan]开始抓取 {len(articles)} 篇正文详情及图片...[/bold cyan]")
                    with Progress(
                        SpinnerColumn(),
                        TextColumn("[progress.description]{task.description}"),
                        BarColumn(),
                        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                        TimeRemainingColumn(),
                        console=console
                    ) as progress:
                        task = progress.add_task("[magenta]抓取正文详情...", total=len(articles))
                        for i, item in enumerate(articles):
                            progress.update(task, description=f"[magenta]正在抓取 ({i+1}/{len(articles)}): {item['title'][:18]}...")
                            await self.crawl_article_content(page, item)
                            progress.advance(task)
                            if self.delay > 0:
                                await asyncio.sleep(self.delay)

                # 第三步：导出数据报表
                console.print("\n[bold cyan]正在导出数据报表...[/bold cyan]")
                excel_path = self.export_excel(articles)
                csv_path = self.export_csv(articles)
                json_path = self.export_json(articles)

                # 输出结果概览
                elapsed = time.time() - start_time
                success_count = sum(1 for a in articles if a.get('status') in ['downloaded', 'cached'])

                table = Table(title="爬取任务完成概览", show_header=True, header_style="bold magenta")
                table.add_column("统计项", style="dim", width=20)
                table.add_column("数值", style="bold green")

                table.add_row("作者名称", self.author_info.get('name', '今日头条作者'))
                table.add_row("发现文章总数", f"{len(articles)} 篇")
                table.add_row("成功下载正文", f"{success_count} 篇")
                table.add_row("Markdown 目录", str(self.articles_dir.resolve()))
                table.add_row("Excel 汇总表", excel_path)
                table.add_row("CSV 汇总表", csv_path)
                table.add_row("JSON 完整数据", json_path)
                table.add_row("总耗时", f"{elapsed:.2f} 秒")

                console.print(table)

            finally:
                await browser.close()
