# -*- coding: utf-8 -*-
"""
今日头条文章采集大师 - 核心采集引擎 (Spider Core)
================================================
- 异步高并发与 Playwright / Patchright 隐蔽内核驱动。
- 支持实时状态流分发 (SSE / Web Event Callback)。
- 支持断点续爬、随时取消任务、图片本地化下载与多格式导出。
"""

import os
import re
import sys
import time
import json
import asyncio
import zipfile
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Callable, Awaitable
import urllib.request

try:
    from patchright.async_api import async_playwright, Browser, BrowserContext, Page
except ImportError:
    from playwright.async_api import async_playwright, Browser, BrowserContext, Page

from bs4 import BeautifulSoup, NavigableString, Tag
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from activation import get_license_status


def sanitize_filename(name: str, max_length: int = 60) -> str:
    """去除文件名中的非法字符并限制最大长度"""
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_length:
        name = name[:max_length]
    return name or 'untitled'


def format_timestamp(ts) -> str:
    if not ts:
        return ''
    try:
        ts = int(ts)
        if ts > 10000000000:
            ts = ts // 1000
        return datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return str(ts)


def format_date_for_file(ts) -> str:
    if not ts:
        return datetime.now().strftime('%Y%m%d')
    try:
        ts = int(ts)
        if ts > 10000000000:
            ts = ts // 1000
        return datetime.fromtimestamp(ts).strftime('%Y%m%d')
    except Exception:
        return datetime.now().strftime('%Y%m%d')


def html_to_markdown(soup_or_tag: Tag, local_image_map: Optional[Dict[str, str]] = None) -> str:
    """将 HTML 转换整理为结构清晰优美的 Markdown"""
    if not soup_or_tag:
        return ''

    def process_node(node):
        if isinstance(node, NavigableString):
            return str(node)
        if not isinstance(node, Tag):
            return ''

        tag_name = node.name.lower()
        if tag_name in ['script', 'style', 'button', 'noscript', 'iframe', 'svg']:
            return ''

        if tag_name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            level = int(tag_name[1])
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n{'#' * level} {inner}\n\n" if inner else ''

        elif tag_name == 'p':
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n{inner}\n\n" if inner else ''

        elif tag_name in ['strong', 'b']:
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"**{inner}**" if inner else ''

        elif tag_name in ['em', 'i']:
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"*{inner}*" if inner else ''

        elif tag_name == 'img':
            src = node.get('data-src') or node.get('src') or ''
            if src.startswith('data:'):
                src = node.get('data-src') or ''
            alt = node.get('alt') or '配图'
            if src:
                final_src = local_image_map.get(src, src) if local_image_map else src
                return f"\n\n![{alt}]({final_src})\n\n"
            return ''

        elif tag_name == 'blockquote':
            inner = ''.join(process_node(c) for c in node.children).strip()
            return f"\n\n> {inner}\n\n" if inner else ''

        elif tag_name in ['ul', 'ol']:
            items = []
            for i, li in enumerate(node.find_all('li', recursive=False)):
                li_text = ''.join(process_node(c) for c in li.children).strip()
                prefix = f"{i+1}. " if tag_name == 'ol' else "- "
                items.append(f"{prefix}{li_text}")
            return '\n\n' + '\n'.join(items) + '\n\n'

        elif tag_name == 'br':
            return '\n'

        elif tag_name == 'a':
            inner = ''.join(process_node(c) for c in node.children).strip()
            href = node.get('href', '')
            if href and inner:
                return f"[{inner}]({href})"
            return inner

        else:
            return ''.join(process_node(c) for c in node.children)

    raw_md = process_node(soup_or_tag)
    cleaned = re.sub(r'\n{3,}', '\n\n', raw_md).strip()
    return cleaned


class ToutiaoSpiderCore:
    """今日头条采集引擎核心"""

    DEFAULT_USER_AGENT = (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/128.0.0.0 Safari/537.36'
    )

    def __init__(
        self,
        author_url: str,
        output_dir: str = './toutiao_output',
        max_articles: Optional[int] = None,
        fetch_content: bool = True,
        download_images: bool = False,
        headless: bool = True,
        delay: float = 0.6,
        scroll_delay: float = 1.0,
        event_callback: Optional[Callable[[str, Dict[str, Any]], Awaitable[None]]] = None
    ):
        self.author_url = author_url.strip()
        self.output_dir = Path(output_dir)
        self.max_articles = max_articles
        self.fetch_content = fetch_content
        self.download_images = download_images
        self.headless = headless
        self.delay = delay
        self.scroll_delay = scroll_delay
        self.event_callback = event_callback

        self.articles_dir = self.output_dir / 'articles'
        self.images_dir = self.output_dir / 'images'
        self.articles_dir.mkdir(parents=True, exist_ok=True)
        if self.download_images:
            self.images_dir.mkdir(parents=True, exist_ok=True)

        self.author_info: Dict[str, Any] = {"url": self.author_url, "name": "今日头条作者"}
        self.articles_map: Dict[str, Dict[str, Any]] = {}
        self.is_cancelled = False
        self.status = "idle"  # idle, running, completed, failed, cancelled
        self.error_msg = ""
        self.start_time = 0.0

    async def emit(self, event_type: str, data: Dict[str, Any]):
        """分发事件给前端/监听者"""
        if self.event_callback:
            try:
                if asyncio.iscoroutinefunction(self.event_callback):
                    await self.event_callback(event_type, data)
                else:
                    self.event_callback(event_type, data)
            except Exception as e:
                print(f"[Emit Error] {e}")

    async def log(self, message: str, level: str = "info"):
        """记录并推送日志"""
        now_str = datetime.now().strftime("%H:%M:%S")
        await self.emit("log", {
            "timestamp": now_str,
            "level": level,
            "message": message
        })

    def cancel(self):
        """取消/中止采集任务"""
        self.is_cancelled = True
        self.status = "cancelled"

    async def _init_browser(self, p):
        # 确保自动发现系统或本地 Playwright/Patchright 浏览器路径
        if not os.environ.get("PLAYWRIGHT_BROWSERS_PATH"):
            local_appdata = Path(os.environ.get("LOCALAPPDATA", "")) / "ms-playwright"
            if local_appdata.exists():
                os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(local_appdata)

        launch_args = [
            '--disable-blink-features=AutomationControlled',
            '--no-sandbox',
            '--disable-dev-shm-usage',
        ]

        browser = None
        last_error = None

        # 1. 尝试默认 chromium 内核
        try:
            browser = await p.chromium.launch(
                headless=self.headless,
                args=launch_args
            )
        except Exception as e:
            last_error = e

        # 2. 尝试系统原生 Microsoft Edge (Windows 10/11 100% 自带，免安装内核)
        if browser is None:
            try:
                browser = await p.chromium.launch(
                    channel="msedge",
                    headless=self.headless,
                    args=launch_args
                )
            except Exception as e:
                last_error = e

        # 3. 尝试系统 Google Chrome
        if browser is None:
            try:
                browser = await p.chromium.launch(
                    channel="chrome",
                    headless=self.headless,
                    args=launch_args
                )
            except Exception as e:
                last_error = e

        # 4. 尝试物理路径检测
        if browser is None:
            candidate_paths = [
                r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            ]
            for c_path in candidate_paths:
                if os.path.exists(c_path):
                    try:
                        browser = await p.chromium.launch(
                            executable_path=c_path,
                            headless=self.headless,
                            args=launch_args
                        )
                        break
                    except Exception as e:
                        last_error = e

        if browser is None:
            raise RuntimeError(f"未能启动任何可用浏览器内核: {last_error}")

        context = await browser.new_context(
            user_agent=self.DEFAULT_USER_AGENT,
            viewport={'width': 1366, 'height': 850},
            locale='zh-CN'
        )
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)
        return browser, context

    async def crawl_author_feed(self, page: Page) -> List[Dict[str, Any]]:
        await self.log(f"正在访问作者主页: {self.author_url}", "info")
        await self.emit("status", {"state": "running", "step": "fetching_list", "progress": 5})

        has_more_flag = True
        captured_batches = 0

        async def on_response(response):
            nonlocal has_more_flag, captured_batches
            if '/api/pc/list/user/feed' in response.url:
                try:
                    data = await response.json()
                    has_more = data.get('has_more', False)
                    items = data.get('data', [])
                    captured_batches += 1

                    new_count = 0
                    for item in items:
                        gid = str(item.get('group_id') or item.get('item_id') or '')
                        if not gid:
                            continue
                        if gid not in self.articles_map:
                            new_count += 1
                            title = (item.get('title') or item.get('abstract') or '无标题').strip()
                            pub_time = item.get('publish_time') or item.get('behot_time')
                            images = []
                            for img in item.get('image_list', []):
                                if isinstance(img, dict) and 'url' in img:
                                    images.append(img['url'])

                            self.articles_map[gid] = {
                                'group_id': gid,
                                'title': title,
                                'abstract': item.get('abstract', ''),
                                'publish_timestamp': pub_time,
                                'publish_time': format_timestamp(pub_time),
                                'read_count': item.get('read_count', 0),
                                'digg_count': item.get('digg_count', 0),
                                'comment_count': item.get('comment_count') or item.get('comments_count', 0),
                                'article_url': f"https://www.toutiao.com/article/{gid}/",
                                'share_url': item.get('share_url', ''),
                                'image_list': images,
                                'has_video': item.get('has_video', False),
                                'content_markdown': '',
                                'local_file': '',
                                'status': 'pending'
                            }

                    if not has_more:
                        has_more_flag = False

                    await self.emit("article_discovered", {
                        "total_count": len(self.articles_map),
                        "batch_new": new_count,
                        "batch_num": captured_batches
                    })
                except Exception:
                    pass

        page.on('response', on_response)

        try:
            await page.goto(self.author_url, wait_until='domcontentloaded', timeout=30000)
            await asyncio.sleep(2)
        except Exception as e:
            await self.log(f"访问作者主页超时或失败: {str(e)}", "error")
            return []

        # 提取作者资料
        try:
            name_el = await page.query_selector('.user-info .name, .name, [class*="user-name"]')
            if name_el:
                self.author_info['name'] = (await name_el.inner_text()).strip()
            desc_el = await page.query_selector('.user-info .desc, .desc, [class*="user-desc"]')
            if desc_el:
                self.author_info['desc'] = (await desc_el.inner_text()).strip()
            avatar_el = await page.query_selector('.user-avatar img, .avatar img, [class*="avatar"] img')
            if avatar_el:
                self.author_info['avatar'] = await avatar_el.get_attribute('src') or ''
        except Exception:
            pass

        await self.log(f"成功识别作者: 【{self.author_info.get('name', '头条创作者')}】", "success")
        await self.emit("author_info", self.author_info)

        # 辅助滚动函数（防止 document.body 为空或跨域 iframe 导致 scrollHeight 报错）
        async def _safe_scroll_down():
            try:
                await page.evaluate("""() => {
                    const h = Math.max(
                        document.body ? document.body.scrollHeight : 0,
                        document.documentElement ? document.documentElement.scrollHeight : 0,
                        document.scrollingElement ? document.scrollingElement.scrollHeight : 0,
                        3000
                    );
                    window.scrollTo(0, h);
                }""")
            except Exception:
                try:
                    await page.keyboard.press("PageDown")
                except Exception:
                    pass

        async def _extract_dom_articles():
            try:
                cards = await page.query_selector_all('a[href*="/article/"], a[href*="/group/"]')
                for card in cards:
                    href = await card.get_attribute('href') or ''
                    m = re.search(r'/(?:article|group)/(\d+)', href)
                    if m:
                        gid = m.group(1)
                        if gid not in self.articles_map:
                            t_text = (await card.inner_text()).strip().split('\n')[0]
                            if t_text and len(t_text) >= 4:
                                self.articles_map[gid] = {
                                    'group_id': gid,
                                    'title': t_text,
                                    'abstract': '',
                                    'publish_timestamp': int(time.time()),
                                    'publish_time': datetime.now().strftime('%Y-%m-%d %H:%M'),
                                    'read_count': 0,
                                    'digg_count': 0,
                                    'comment_count': 0,
                                    'article_url': f"https://www.toutiao.com/article/{gid}/",
                                    'share_url': f"https://www.toutiao.com/article/{gid}/",
                                    'image_list': [],
                                    'has_video': False,
                                    'content_markdown': '',
                                    'local_file': '',
                                    'status': 'pending'
                                }
            except Exception:
                pass

        # 滚动加载
        last_count = 0
        same_count_streak = 0

        while has_more_flag and not self.is_cancelled:
            await _safe_scroll_down()
            await asyncio.sleep(self.scroll_delay)
            await _extract_dom_articles()

            current_count = len(self.articles_map)
            await self.log(f"正在滚动拉取文章列表，已发现 {current_count} 篇...", "info")

            if self.max_articles and current_count >= self.max_articles:
                await self.log(f"已达到目标最大采集篇数 ({self.max_articles} 篇)", "info")
                break

            if current_count == last_count:
                same_count_streak += 1
                if same_count_streak >= 5:
                    try:
                        await page.evaluate('window.scrollBy(0, -400)')
                    except Exception:
                        pass
                    await asyncio.sleep(0.5)
                    await _safe_scroll_down()
                    await asyncio.sleep(self.scroll_delay)
                    await _extract_dom_articles()
                    if len(self.articles_map) == last_count:
                        await self.log("已滚动到页面底部，所有文章列表加载完毕。", "info")
                        break
            else:
                same_count_streak = 0
                last_count = current_count

        article_list = list(self.articles_map.values())
        if self.max_articles:
            article_list = article_list[:self.max_articles]

        await self.log(f"文章列表获取完毕，共计发现 {len(article_list)} 篇有效文章", "success")
        return article_list

    def _download_image(self, img_url: str, save_path: Path) -> bool:
        try:
            req = urllib.request.Request(
                img_url,
                headers={
                    'User-Agent': self.DEFAULT_USER_AGENT,
                    'Referer': 'https://www.toutiao.com/'
                }
            )
            with urllib.request.urlopen(req, timeout=10) as resp, open(save_path, 'wb') as f:
                f.write(resp.read())
            return True
        except Exception:
            return False

    async def crawl_article_content(self, page: Page, item: Dict[str, Any], index: int, total: int) -> Dict[str, Any]:
        if self.is_cancelled:
            item['status'] = 'cancelled'
            return item

        gid = item['group_id']
        url = item['article_url']
        pub_date = format_date_for_file(item['publish_timestamp'])
        safe_title = sanitize_filename(item['title'])
        filename = f"{pub_date}_{gid}_{safe_title}.md"
        filepath = self.articles_dir / filename

        # 断点续爬检查
        if filepath.exists() and filepath.stat().st_size > 100:
            item['local_file'] = str(filepath.name)
            item['status'] = 'cached'
            await self.log(f"[{index}/{total}] 已存在本地缓存: {item['title'][:20]}...", "info")
            await self.emit("article_processed", {"index": index, "total": total, "item": item})
            return item

        try:
            await page.goto(url, wait_until='domcontentloaded', timeout=20000)

            article_node = None
            for selector in ['article', '.article-content', '.tt-article-content', '.s-content', '.main-content']:
                try:
                    await page.wait_for_selector(selector, timeout=3000)
                    article_node = selector
                    break
                except Exception:
                    continue

            if not article_node:
                html = await page.content()
                soup = BeautifulSoup(html, 'html.parser')
                article_tag = soup.find('article') or soup.find('div', class_='article-content')
            else:
                article_html = await page.inner_html(article_node)
                soup = BeautifulSoup(article_html, 'html.parser')
                article_tag = soup

            try:
                h1_el = await page.query_selector('h1')
                if h1_el:
                    page_title = (await h1_el.inner_text()).strip()
                    if page_title:
                        item['title'] = page_title
            except Exception:
                pass

            # 配图下载
            local_image_map = {}
            if self.download_images and article_tag:
                img_nodes = article_tag.find_all('img')
                if img_nodes:
                    article_img_dir = self.images_dir / gid
                    article_img_dir.mkdir(parents=True, exist_ok=True)
                    for idx, img_tag in enumerate(img_nodes, 1):
                        src = img_tag.get('data-src') or img_tag.get('src') or ''
                        if src and not src.startswith('data:'):
                            ext = 'jpg'
                            if '.png' in src: ext = 'png'
                            elif '.webp' in src: ext = 'webp'
                            elif '.gif' in src: ext = 'gif'
                            img_filename = f"{idx:02d}.{ext}"
                            img_save_path = article_img_dir / img_filename
                            if not img_save_path.exists():
                                self._download_image(src, img_save_path)
                            rel_img_path = f"../images/{gid}/{img_filename}"
                            local_image_map[src] = rel_img_path

            content_md = html_to_markdown(article_tag, local_image_map)
            item['content_markdown'] = content_md

            author_name = self.author_info.get('name', '今日头条作者')
            md_full_text = (
                f"---\n"
                f"title: \"{item['title']}\"\n"
                f"author: \"{author_name}\"\n"
                f"publish_time: \"{item['publish_time']}\"\n"
                f"group_id: \"{gid}\"\n"
                f"read_count: {item.get('read_count', 0)}\n"
                f"digg_count: {item.get('digg_count', 0)}\n"
                f"comment_count: {item.get('comment_count', 0)}\n"
                f"source_url: \"{url}\"\n"
                f"---\n\n"
                f"# {item['title']}\n\n"
                f"> **发布时间**：{item['publish_time']} | **阅读量**：{item.get('read_count', 0)} | **点赞**：{item.get('digg_count', 0)} | **原文链接**：[{url}]({url})\n\n"
                f"---\n\n"
                f"{content_md}\n"
            )

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(md_full_text)

            item['local_file'] = str(filepath.name)
            item['status'] = 'downloaded'
            await self.log(f"[{index}/{total}] 抓取完成: {item['title'][:20]}...", "success")

        except Exception as e:
            item['status'] = f"failed: {str(e)}"
            await self.log(f"[{index}/{total}] 抓取失败 ({item['title'][:15]}): {str(e)}", "warning")

        await self.emit("article_processed", {"index": index, "total": total, "item": item})
        return item

    def export_excel(self, articles: List[Dict[str, Any]]) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文章汇总"

        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Microsoft YaHei", size=10)
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        headers = ["序号", "文章ID", "发布时间", "文章标题", "阅读量", "点赞量", "评论量", "文章摘要", "原文链接", "本地Markdown文件"]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        for idx, item in enumerate(articles, 1):
            row_data = [
                idx,
                item.get('group_id', ''),
                item.get('publish_time', ''),
                item.get('title', ''),
                item.get('read_count', 0),
                item.get('digg_count', 0),
                item.get('comment_count', 0),
                item.get('abstract', ''),
                item.get('article_url', ''),
                item.get('local_file', '')
            ]
            ws.append(row_data)
            row_num = idx + 1
            ws.row_dimensions[row_num].height = 22

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = border
                if col_idx in [1, 2, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [3, 4, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        col_widths = {1: 8, 2: 24, 3: 20, 4: 45, 5: 12, 6: 12, 7: 12, 8: 40, 9: 35, 10: 35}
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        excel_path = self.output_dir / 'articles_summary.xlsx'
        wb.save(excel_path)
        return str(excel_path)

    def export_csv(self, articles: List[Dict[str, Any]]) -> str:
        import csv
        csv_path = self.output_dir / 'articles_summary.csv'
        headers = ["序号", "文章ID", "发布时间", "文章标题", "阅读量", "点赞量", "评论量", "摘要", "原文链接", "本地文件"]
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for idx, item in enumerate(articles, 1):
                writer.writerow([
                    idx,
                    item.get('group_id', ''),
                    item.get('publish_time', ''),
                    item.get('title', ''),
                    item.get('read_count', 0),
                    item.get('digg_count', 0),
                    item.get('comment_count', 0),
                    item.get('abstract', ''),
                    item.get('article_url', ''),
                    item.get('local_file', '')
                ])
        return str(csv_path)

    def export_json(self, articles: List[Dict[str, Any]]) -> str:
        json_path = self.output_dir / 'articles_data.json'
        payload = {
            'crawler_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'author_info': self.author_info,
            'total_articles': len(articles),
            'articles': articles
        }
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return str(json_path)

    def export_zip(self) -> str:
        """将输出目录打包为 zip 文件"""
        zip_path = self.output_dir / 'articles_package.zip'
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(self.output_dir):
                for file in files:
                    full_p = Path(root) / file
                    if full_p == zip_path:
                        continue
                    rel_p = full_p.relative_to(self.output_dir)
                    zf.write(full_p, rel_p)
        return str(zip_path)

    async def run(self) -> Dict[str, Any]:
        """主执行逻辑"""
        self.start_time = time.time()
        self.status = "running"
        self.is_cancelled = False

        # 检查 VIP 会员权限并进行配额限制
        license_info = get_license_status()
        if not license_info["is_vip"]:
            max_limit = license_info["max_articles_per_crawl"]
            if self.max_articles is None or self.max_articles > max_limit:
                await self.log(f"【体验版限制】单次采集最多允许 {max_limit} 篇（VIP 无限制）", "warning")
                self.max_articles = max_limit
            if self.download_images:
                await self.log("【体验版限制】已自动关闭高清插图批量下载功能（VIP 专享）", "warning")
                self.download_images = False

        await self.log(f"采集任务启动: 目标作者 [{self.author_url}]", "info")
        await self.emit("status", {"state": "running", "step": "starting", "progress": 0})

        articles: List[Dict[str, Any]] = []

        try:
            async with async_playwright() as p:
                browser, context = await self._init_browser(p)
                page = await context.new_page()

                try:
                    # 1. 抓取文章列表
                    articles = await self.crawl_author_feed(page)
                    if self.is_cancelled:
                        await self.log("任务已由用户手动中止", "warning")
                        self.status = "cancelled"
                        return {"status": "cancelled", "articles": []}

                    if not articles:
                        await self.log("未获取到任何文章，请检查作者主页链接", "error")
                        self.status = "failed"
                        return {"status": "failed", "articles": []}

                    # 2. 抓取正文详情
                    if self.fetch_content:
                        total_articles = len(articles)
                        await self.log(f"开始抓取 {total_articles} 篇正文详情及排版...", "info")
                        for idx, item in enumerate(articles, 1):
                            if self.is_cancelled:
                                await self.log("任务已由用户手动中止", "warning")
                                break
                            progress_pct = int(10 + (idx / total_articles) * 80)
                            await self.emit("status", {
                                "state": "running",
                                "step": "fetching_details",
                                "progress": progress_pct,
                                "current_index": idx,
                                "total": total_articles,
                                "current_title": item['title']
                            })
                            await self.crawl_article_content(page, item, idx, total_articles)
                            if self.delay > 0:
                                await asyncio.sleep(self.delay)

                    # 3. 导出报表
                    await self.log("正在生成汇总报表 (Excel / CSV / JSON)...", "info")
                    excel_path = self.export_excel(articles)
                    csv_path = self.export_csv(articles)
                    json_path = self.export_json(articles)
                    zip_path = self.export_zip()

                    elapsed = time.time() - self.start_time
                    success_count = sum(1 for a in articles if a.get('status') in ['downloaded', 'cached'])

                    summary = {
                        "author_name": self.author_info.get('name', '今日头条作者'),
                        "total_count": len(articles),
                        "success_count": success_count,
                        "elapsed_seconds": round(elapsed, 2),
                        "excel_path": excel_path,
                        "csv_path": csv_path,
                        "json_path": json_path,
                        "zip_path": zip_path,
                        "output_dir": str(self.output_dir.resolve())
                    }

                    if self.is_cancelled:
                        self.status = "cancelled"
                        await self.log(f"任务已由用户中止！已导出截至目前抓取的 {success_count} 篇文章数据", "warning")
                        await self.emit("status", {"state": "cancelled", "step": "cancelled", "progress": progress_pct if 'progress_pct' in locals() else 0})
                        await self.emit("export_ready", summary)
                        return {
                            "status": "cancelled",
                            "summary": summary,
                            "articles": articles
                        }

                    self.status = "completed"
                    await self.log(f"任务圆满完成！成功处理 {success_count}/{len(articles)} 篇文章，耗时 {elapsed:.1f} 秒", "success")
                    await self.emit("status", {"state": "completed", "step": "done", "progress": 100})
                    await self.emit("export_ready", summary)

                    return {
                        "status": "completed",
                        "summary": summary,
                        "articles": articles
                    }

                finally:
                    await browser.close()

        except Exception as e:
            self.status = "failed"
            self.error_msg = str(e)
            await self.log(f"采集过程出现异常: {str(e)}", "error")
            await self.emit("status", {"state": "failed", "step": "error", "error": str(e)})
            return {
                "status": "failed",
                "error": str(e),
                "articles": articles
            }
